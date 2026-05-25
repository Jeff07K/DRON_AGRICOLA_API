"""
export.py — Dron Agrícola MPU6050
Jeffrey Bejarano — 67001609 · Universidad Católica de Colombia · 2026-1

Genera un Excel profesional con:
  Hoja 1 · Datos          — tabla de datos crudos formateada, colores alternos
  Hoja 2 · Estadística    — tabla completa + gráfico de barras con fórmulas Excel
  Hoja 3 · Frecuencias    — histogramas con bins reales + gráficos de columnas
  Hoja 4 · Correlación    — matriz Pearson coloreada + regresión lineal
  Hoja 5 · Probabilístico — Normal, TCL, IC 95%, Binomial, Geométrica
  Hoja 6 · Gráficas       — gráficos de línea de los 6 ejes por disparo
"""

import io, math
from typing import List

import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, ScatterChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList

from models import SensorData

# ─── Paleta ───────────────────────────────────────────────────────────────────
C = {
    'dark_green':  '1A4731',  # header fondo
    'mid_green':   '276749',  # sub-header
    'light_green': 'E8F5E9',  # fila alterna 1
    'white':       'FFFFFF',  # fila alterna 2
    'accent_blue': '1565C0',  # valores numéricos destacados
    'accent_amber':'F57F17',  # advertencias
    'accent_red':  'B71C1C',  # valores negativos extremos
    'positive':    '1B5E20',  # texto valor positivo
    'negative':    'B71C1C',  # texto valor negativo
    'muted':       '607D8B',  # texto secundario
    'gold':        'F9A825',  # títulos de sección
    'pearl':       'F5F5F5',  # fondos neutrales
    'corr_high':   'C8E6C9',  # correlación fuerte
    'corr_mid':    'FFF9C4',  # correlación media
    'corr_low':    'FFCDD2',  # correlación baja/nula
    'sig':         'E8F5E9',  # p < 0.05
    'nsig':        'FFCDD2',  # p >= 0.05
}

THIN  = Side(border_style='thin',   color='BDBDBD')
THICK = Side(border_style='medium', color='1A4731')
B_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
B_BOT = Border(bottom=THICK)

def fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def font(bold=False, color='000000', size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic,
                name='Calibri')

def align(h='center', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def cell(ws, row, col, value='', bold=False, bg=None, fg='000000',
         size=10, h='left', v='center', wrap=False, italic=False,
         border=True, num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = font(bold=bold, color=fg, size=size, italic=italic)
    c.alignment = align(h=h, v=v, wrap=wrap)
    if bg:  c.fill   = fill(bg)
    if border: c.border = B_ALL
    if num_fmt: c.number_format = num_fmt
    return c

def hdr(ws, row, col, text, span=1):
    """Celda de encabezado principal — verde oscuro, texto blanco, bold"""
    c = cell(ws, row, col, text, bold=True, bg=C['dark_green'],
             fg='FFFFFF', h='center', size=10)
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+span-1)
    return c

def sub_hdr(ws, row, col, text, span=1):
    """Sub-encabezado — verde medio"""
    c = cell(ws, row, col, text, bold=True, bg=C['mid_green'],
             fg='FFFFFF', h='center', size=9)
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+span-1)
    return c

def section_title(ws, row, col, text, span=8):
    """Título de sección — dorado"""
    c = cell(ws, row, col, text, bold=True, bg=C['dark_green'],
             fg=C['gold'], h='left', size=12, border=False)
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row, end_column=col+span-1)
    ws.row_dimensions[row].height = 22
    return c

def autowidth(ws, min_w=8, max_w=40):
    for col in ws.columns:
        w = max((len(str(c.value or '')) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = \
            min(max(w + 3, min_w), max_w)

# ─── Estadísticas ─────────────────────────────────────────────────────────────
def _stat(values):
    n = len(values)
    if n < 1: return {}
    s  = sorted(values)
    mu = sum(s) / n
    va = sum((v-mu)**2 for v in s) / n
    sd = math.sqrt(va)
    def pct(p):
        i = (n-1)*p; lo,hi = int(i), math.ceil(i)
        return s[lo] if lo==hi else s[lo]+(s[hi]-s[lo])*(i-lo)
    q1,med,q3 = pct(.25),pct(.5),pct(.75)
    iqr = q3-q1
    cv  = sd/abs(mu)*100 if abs(mu)>1e-9 else float('inf')
    return dict(n=n, mean=round(mu,4), var=round(va,4), std=round(sd,4),
                min=round(s[0],4), q1=round(q1,4), med=round(med,4),
                q3=round(q3,4), max=round(s[-1],4), iqr=round(iqr,4),
                cv=round(cv,2) if math.isfinite(cv) else 9999)

def _pearson(xs, ys):
    n = len(xs)
    if n < 3: return 0, 1   # necesita al menos 3 puntos
    mx,my = sum(xs)/n, sum(ys)/n
    num = sum((x-mx)*(y-my) for x,y in zip(xs,ys))
    dx  = math.sqrt(sum((x-mx)**2 for x in xs))
    dy  = math.sqrt(sum((y-my)**2 for y in ys))
    if dx*dy == 0: return 0, 1
    r   = max(-1, min(1, num/(dx*dy)))
    t   = r*math.sqrt(n-2)/math.sqrt(max(1e-10,1-r*r))
    p   = min(1, math.exp(-0.717*abs(t) - 0.416*t*t))
    return round(r,4), round(p,4)

def _linreg(xs, ys):
    n = len(xs)
    if n < 2: return 0,0,0
    mx,my = sum(xs)/n, sum(ys)/n
    b1n = sum((x-mx)*(y-my) for x,y in zip(xs,ys))
    b1d = sum((x-mx)**2 for x in xs)
    if b1d == 0: return round(my,4), 0, 0
    b1  = b1n/b1d; b0 = my-b1*mx
    sst = sum((y-my)**2 for y in ys)
    sse = sum((y-(b0+b1*x))**2 for x,y in zip(xs,ys))
    r2  = 1-sse/sst if sst>0 else 0
    return round(b0,4), round(b1,4), round(r2,4)

def _norm_cdf(x, mu, sigma):
    if sigma == 0: return (0 if x < mu else 1)
    z = (x-mu)/(sigma*math.sqrt(2))
    t = 1/(1+0.3275911*abs(z))
    e = 1-(((((1.061405429*t-1.453152027)*t+1.421413741)*t
              -0.284496736)*t+0.254829592)*t)*math.exp(-z*z)
    return 0.5*(1+(e if z>=0 else -e))

def _freq(values, bins):
    """bins = [(lo,hi), ...]. Último bin incluye el extremo."""
    counts = []
    for i,(lo,hi) in enumerate(bins):
        if i == len(bins)-1:
            counts.append(sum(1 for v in values if lo<=v<=hi))
        else:
            counts.append(sum(1 for v in values if lo<=v<hi))
    return counts

# ─── MAIN ────────────────────────────────────────────────────────────────────
def generate_excel(records: List[SensorData]) -> bytes:

    # Filtrar ceros (sensor offline)
    valid = [r for r in records if not (
        r.accel_x==0 and r.accel_y==0 and r.accel_z==0 and
        r.gyro_x==0  and r.gyro_y==0  and r.gyro_z==0)]

    # Agrupar por disparo para gráficas
    by_d = {}
    for r in valid:
        by_d.setdefault(r.disparo, []).append(r)
    d_ids = sorted(by_d.keys())

    AX = ['Accel X','Accel Y','Accel Z','Gyro X','Gyro Y','Gyro Z']
    UNITS = {'Accel X':'g','Accel Y':'g','Accel Z':'g',
             'Gyro X':'°/s','Gyro Y':'°/s','Gyro Z':'°/s'}
    COLORS_AX = ['4472C4','7030A0','70AD47','FF0000','FFC000','FF6600']

    def ax_vals(name):
        key = name.lower().replace(' ','_')
        return [getattr(r, key) for r in valid]

    wb = openpyxl.Workbook()

    # ═══════════════════════════════════════════════════════════════════════
    # HOJA 1 — DATOS SENSOR
    # ═══════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = 'Datos'
    ws1.freeze_panes = 'A3'

    # Título
    section_title(ws1, 1, 1, '  DRON AGRÍCOLA · MPU6050 — Datos del Sensor', span=9)

    # Headers
    hdrs = ['N°','Timestamp','Disparo','Accel X (g)','Accel Y (g)',
            'Accel Z (g)','Gyro X (°/s)','Gyro Y (°/s)','Gyro Z (°/s)']
    for ci, h in enumerate(hdrs, 1):
        hdr(ws1, 2, ci, h)

    # Data rows
    for ri, r in enumerate(valid, 3):
        bg = C['light_green'] if ri % 2 == 0 else C['white']
        vals_row = [ri-2, str(r.timestamp)[:19], r.disparo,
                    r.accel_x, r.accel_y, r.accel_z,
                    r.gyro_x,  r.gyro_y,  r.gyro_z]
        for ci, v in enumerate(vals_row, 1):
            fg = '000000'
            if ci == 3: fg = C['accent_blue']  # disparo
            if ci >= 4:
                fg = C['positive'] if isinstance(v,(int,float)) and v>=0 \
                     else C['negative']
            nf = '0.0000' if ci >= 4 else None
            cell(ws1, ri, ci, v, bg=bg, fg=fg, h='center',
                 num_fmt=nf)

    # Totals row
    n_data = len(valid)
    last_r = 2 + n_data
    cell(ws1, last_r+1, 1, f'Total: {n_data} registros válidos',
         bold=True, bg=C['dark_green'], fg='FFFFFF', h='center')
    ws1.merge_cells(f'A{last_r+1}:I{last_r+1}')
    # Apply fill/font to merged cells
    for col in range(2, 10):
        c2 = ws1.cell(row=last_r+1, column=col)
        c2.fill = fill(C['dark_green'])
        c2.border = B_ALL

    ws1.column_dimensions['A'].width = 5
    ws1.column_dimensions['B'].width = 22
    ws1.column_dimensions['C'].width = 10
    for col in 'DEFGHI':
        ws1.column_dimensions[col].width = 14

    # ═══════════════════════════════════════════════════════════════════════
    # HOJA 2 — ESTADÍSTICA DESCRIPTIVA
    # ═══════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet('Estadística')

    section_title(ws2, 1, 1,
        f'  ESTADÍSTICA DESCRIPTIVA — n = {len(valid)} registros válidos', span=14)

    # Advertencia si n < 10
    if len(valid) < 10:
        c2 = cell(ws2, 2, 1,
            f'⚠ ADVERTENCIA: Solo {len(valid)} registros. '
            'Agrega más disparos para resultados estadísticos confiables. '
            'Se necesitan al menos 30 registros (TCL).',
            bold=True, bg='FFF9C4', fg=C['accent_amber'], wrap=True,
            h='left', border=False)
        ws2.merge_cells('A2:N2')
        ws2.row_dimensions[2].height = 28
        start_row = 4
    else:
        start_row = 3

    stat_cols = ['Variable','Unidad','N','Media','Varianza','Desv. Est.',
                 'Mínimo','Q1 (25%)','Mediana','Q3 (75%)','Máximo',
                 'IQR','CV (%)','Estabilidad']
    for ci, h in enumerate(stat_cols, 1):
        hdr(ws2, start_row, ci, h)

    for ri, name in enumerate(AX, start_row+1):
        vals = ax_vals(name)
        s    = _stat(vals)
        bg   = C['light_green'] if ri % 2 == 0 else C['white']
        cv   = s.get('cv', 9999)
        if   cv < 30:   stab = '✓ Estable';    sfg = C['positive']
        elif cv < 100:  stab = '▲ Moderada';   sfg = C['accent_amber']
        elif cv < 300:  stab = '⚠ Alta';       sfg = C['accent_red']
        else:           stab = '✕ Muy alta';   sfg = C['accent_red']
        cv_txt = f'{cv:.2f}%' if cv < 9999 else '∞'

        row_vals = [name, UNITS[name], s.get('n',0),
                    s.get('mean',0), s.get('var',0), s.get('std',0),
                    s.get('min',0), s.get('q1',0), s.get('med',0),
                    s.get('q3',0), s.get('max',0), s.get('iqr',0),
                    cv_txt, stab]

        for ci, v in enumerate(row_vals, 1):
            fg = sfg if ci == 14 else \
                 (COLORS_AX[ri-start_row-1] if ci == 1 else '000000')
            bold = ci in (1, 4, 14)
            nf   = '0.0000' if ci in range(4,13) else None
            cell(ws2, ri, ci, v, bold=bold, bg=bg, fg=fg,
                 h='center', num_fmt=nf)

    # Gráfico: Media de acelerómetro
    chart_a = BarChart()
    chart_a.type = 'col'; chart_a.title = 'Media Acelerómetro (g)'
    chart_a.y_axis.title = 'g'; chart_a.shape = 4
    data_r   = Reference(ws2, min_col=4, min_row=start_row,
                         max_row=start_row+3)
    labels_r = Reference(ws2, min_col=1, min_row=start_row+1,
                         max_row=start_row+3)
    chart_a.add_data(data_r, titles_from_data=True)
    chart_a.set_categories(labels_r)
    chart_a.width = 14; chart_a.height = 10
    ws2.add_chart(chart_a, 'P3')

    # Gráfico: Media giroscopio
    chart_g = BarChart()
    chart_g.type = 'col'; chart_g.title = 'Media Giroscopio (°/s)'
    chart_g.y_axis.title = '°/s'; chart_g.shape = 4
    # Gyro rows: header at start_row, data at start_row+4 to start_row+6
    data_g   = Reference(ws2, min_col=4, min_row=start_row,
                         max_row=start_row+6)
    labels_g = Reference(ws2, min_col=1, min_row=start_row+4,
                         max_row=start_row+6)
    chart_g.add_data(data_g, titles_from_data=True)
    chart_g.set_categories(labels_g)
    chart_g.width = 14; chart_g.height = 10
    ws2.add_chart(chart_g, 'P18')

    # Nota
    note_row = start_row + 8
    cell(ws2, note_row, 1,
         'Nota: CV = Desv.Est / |Media| × 100  ·  IQR = Q3 − Q1  '
         '·  ✓ CV<30% = estable  ·  ▲ 30–100%  ·  ⚠ 100–300%  ·  ✕ >300%',
         italic=True, fg=C['muted'], h='left', border=False, wrap=True)
    ws2.merge_cells(f'A{note_row}:N{note_row}')

    autowidth(ws2)

    # ═══════════════════════════════════════════════════════════════════════
    # HOJA 3 — DISTRIBUCIÓN DE FRECUENCIAS
    # ═══════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet('Frecuencias')
    section_title(ws3, 1, 1,
        f'  DISTRIBUCIÓN DE FRECUENCIAS — bins ajustados al rango real', span=12)

    if len(valid) < 3:
        cell(ws3, 2, 1,
             f'⚠ Solo {len(valid)} registros — histogramas no representativos. '
             'Genera más disparos para distribuciones significativas.',
             bold=True, bg='FFF9C4', fg=C['accent_amber'], wrap=True,
             h='left', border=False)
        ws3.merge_cells('A2:L2')
        ws3.row_dimensions[2].height = 28

    BINS = {
        'Accel X': [(-0.9,-0.6),(-0.6,-0.3),(-0.3,0.0),(0.0,0.3),(0.3,0.6),(0.6,0.9)],
        'Accel Y': [(-1.0,-0.6),(-0.6,-0.3),(-0.3,0.0),(0.0,0.3),(0.3,0.6),(0.6,1.0)],
        'Accel Z': [(-0.7,0.0),(0.0,0.4),(0.4,0.7),(0.7,0.9),(0.9,1.1),(1.1,1.4)],
        'Gyro X':  [(-100,-50),(-50,-20),(-20,-5),(-5,5),(5,20),(20,100)],
        'Gyro Y':  [(-40,-20),(-20,-5),(-5,5),(5,20),(20,40),(40,72)],
        'Gyro Z':  [(-55,-20),(-20,-5),(-5,5),(5,20),(20,50),(50,110)],
    }
    HEALTHY = {  # bins que representan operación normal
        'Accel Z': [(0.9,1.1)],
        'Gyro X':  [(-20,-5),(-5,5)],
    }

    col_start_map = [1, 5, 9]   # columnas de inicio para las 3 primeras variables
    for group_start_row, group in enumerate([[('Accel X',0),('Accel Y',4),('Accel Z',8)],
                                              [('Gyro X',0),('Gyro Y',4),('Gyro Z',8)]]):
        base_row = 4 + group_start_row * 11
        for name, col_off in group:
            col = col_off + 1
            vals = ax_vals(name)
            bins = BINS[name]
            cnts = _freq(vals, bins)
            total = sum(cnts) or 1

            # Variable title
            sub_hdr(ws3, base_row, col,
                    f'{name} ({UNITS[name]})', span=3)

            # Column headers
            for ci, h in enumerate(['Intervalo','Frec. Abs.','Frec. Rel.%'], col):
                hdr(ws3, base_row+1, ci, h)

            for bi, ((lo,hi), cnt) in enumerate(zip(bins, cnts)):
                r = base_row + 2 + bi
                bg = C['light_green'] if bi%2==0 else C['white']
                # Verde especial para bins "saludables"
                is_healthy = name in HEALTHY and (lo,hi) in HEALTHY[name]
                if is_healthy: bg = 'C8E6C9'

                lbl = f'{lo} a {hi}'
                pct = round(cnt/total*100, 1)
                cell(ws3, r, col,   lbl, bg=bg, h='center')
                cell(ws3, r, col+1, cnt, bg=bg, h='center',
                     fg=C['accent_blue'] if cnt>0 else C['muted'])
                cell(ws3, r, col+2, pct, bg=bg, h='center',
                     num_fmt='0.0"%"')

            # Mini gráfico de barras para cada variable
            bar = BarChart()
            bar.type  = 'col'
            bar.title = f'Histograma {name} ({UNITS[name]})'
            bar.y_axis.title = 'Frecuencia'
            bar.x_axis.title = f'Intervalo ({UNITS[name]})'
            bar.shape = 4; bar.width = 12; bar.height = 9
            data_ref  = Reference(ws3, min_col=col+1,
                                  min_row=base_row+1,
                                  max_row=base_row+1+len(bins))
            label_ref = Reference(ws3, min_col=col,
                                  min_row=base_row+2,
                                  max_row=base_row+1+len(bins))
            bar.add_data(data_ref, titles_from_data=True)
            bar.set_categories(label_ref)
            # Posicionar gráfico debajo de la tabla
            chart_col = get_column_letter(col)
            chart_row = base_row + 2 + len(bins) + 1
            ws3.add_chart(bar, f'{chart_col}{chart_row}')

    autowidth(ws3)

    # ═══════════════════════════════════════════════════════════════════════
    # HOJA 4 — CORRELACIÓN Y REGRESIÓN
    # ═══════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet('Correlación')
    section_title(ws4, 1, 1,
        '  CORRELACIÓN DE PEARSON + REGRESIÓN LINEAL', span=14)

    if len(valid) < 3:
        cell(ws4, 2, 1,
             f'⚠ Solo {len(valid)} registros — se necesitan ≥ 3 para calcular '
             'correlación. Todos los valores serán 0. Genera más disparos.',
             bold=True, bg='FFF9C4', fg=C['accent_amber'], wrap=True,
             h='left', border=False)
        ws4.merge_cells('A2:N2')
        ws4.row_dimensions[2].height = 28
        sr = 4
    else:
        sr = 3

    # ── Matriz r de Pearson ──
    sub_hdr(ws4, sr, 1, 'Matriz r (Pearson)', span=7)
    for ci, name in enumerate(AX, 2):
        hdr(ws4, sr+1, ci, name)
    for ri, name in enumerate(AX, sr+2):
        hdr(ws4, ri, 1, name)

    for ri, na in enumerate(AX, sr+2):
        for ci, nb in enumerate(AX, 2):
            r, p = _pearson(ax_vals(na), ax_vals(nb))
            if na == nb:
                bg = 'D5E8D4'  # diagonal
            elif abs(r) > 0.7:
                bg = C['corr_high']
            elif abs(r) > 0.4:
                bg = C['corr_mid']
            else:
                bg = C['corr_low']
            bold = na == nb
            cell(ws4, ri, ci, r if na!=nb else 1.0,
                 bold=bold, bg=bg, h='center', num_fmt='0.0000')

    # ── Matriz p-values ──
    sub_hdr(ws4, sr, 9, 'Matriz p-value', span=6)
    for ci, name in enumerate(AX, 9):
        hdr(ws4, sr+1, ci, name)
    for ri, name in enumerate(AX, sr+2):
        hdr(ws4, ri, 9, name)

    for ri, na in enumerate(AX, sr+2):
        for ci, nb in enumerate(AX, 9):
            r, p = _pearson(ax_vals(na), ax_vals(nb))
            if na == nb:
                bg='D5E8D4'; val=0.0
            else:
                bg = C['sig'] if p < 0.05 else C['nsig']
                val = p
            cell(ws4, ri, ci, val, bg=bg, h='center', num_fmt='0.0000')

    # Leyenda
    ley_row = sr + 9
    for col, (bg, txt) in enumerate([
        ('D5E8D4','Diagonal — r = 1 (variable consigo misma)'),
        (C['corr_high'],'|r| > 0.7 — Correlación fuerte'),
        (C['corr_mid'], '|r| > 0.4 — Correlación moderada'),
        (C['corr_low'], '|r| ≤ 0.4 — Correlación débil / independiente'),
        (C['sig'],      'p < 0.05 — Estadísticamente significativo'),
        (C['nsig'],     'p ≥ 0.05 — No significativo'),
    ], 1):
        cell(ws4, ley_row, col, txt, bg=bg, h='left',
             italic=True, fg=C['muted'], size=9)

    # ── Regresión lineal ──
    reg_row = ley_row + 2
    section_title(ws4, reg_row, 1,
        '  REGRESIÓN LINEAL — Ŷ = B0 + B1·X', span=8)
    reg_hdrs = ['Variable X','Variable Y','B0 (intercepto)',
                'B1 (pendiente)','R²','r (Pearson)','p-value','Interpretación']
    for ci, h in enumerate(reg_hdrs, 1):
        hdr(ws4, reg_row+1, ci, h)

    pairs = [('Accel X','Accel Y'),('Gyro X','Gyro Z'),('Accel Z','Gyro X')]
    for ri, (xa, ya) in enumerate(pairs, reg_row+2):
        b0, b1, r2 = _linreg(ax_vals(xa), ax_vals(ya))
        r, p       = _pearson(ax_vals(xa), ax_vals(ya))
        bg = C['light_green'] if ri%2==0 else C['white']
        if   r2 > 0.7: interp = f'Fuerte: {r2*100:.1f}% de {ya} explicado por {xa}'
        elif r2 > 0.3: interp = f'Moderada: {r2*100:.1f}% explicado'
        else:          interp = f'Débil: {r2*100:.1f}% — ejes mayormente independientes'
        for ci, v in enumerate([xa,ya,b0,b1,r2,r,p,interp], 1):
            nf = '0.0000' if ci in (3,4,5,6,7) else None
            cell(ws4, ri, ci, v, bg=bg, h='center', num_fmt=nf,
                 wrap=(ci==8))

    autowidth(ws4)

    # ═══════════════════════════════════════════════════════════════════════
    # HOJA 5 — ANÁLISIS PROBABILÍSTICO
    # ═══════════════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet('Probabilístico')
    section_title(ws5, 1, 1,
        '  ANÁLISIS PROBABILÍSTICO — Normal, TCL, IC 95%, Binomial, Geométrica',
        span=9)

    if len(valid) < 3:
        cell(ws5, 2, 1,
             f'⚠ Solo {len(valid)} registros. Los modelos probabilísticos '
             'requieren al menos 30 datos para ser válidos (TCL). '
             'Los valores mostrados son ilustrativos.',
             bold=True, bg='FFF9C4', fg=C['accent_amber'], wrap=True,
             h='left', border=False)
        ws5.merge_cells('A2:I2')
        ws5.row_dimensions[2].height = 35
        sr5 = 4
    else:
        sr5 = 3

    # ── Distribución Normal ──
    sub_hdr(ws5, sr5, 1, 'Distribución Normal — Accel X ~ N(μ, σ²)', span=9)
    norm_hdrs = ['Variable','Distribución','μ','σ²','σ',
                 'P(X<0)','P(X>0)','IC 95% Inf','IC 95% Sup']
    for ci, h in enumerate(norm_hdrs, 1):
        hdr(ws5, sr5+1, ci, h)

    for ri, name in enumerate(AX, sr5+2):
        vals = ax_vals(name)
        s    = _stat(vals)
        mu, sigma = s.get('mean',0), s.get('std',0)
        p_neg = round(_norm_cdf(0, mu, sigma), 4)
        p_pos = round(1-p_neg, 4)
        n     = s.get('n',1) or 1
        se    = sigma/math.sqrt(n)
        ic_lo = round(mu - 1.96*se, 4)
        ic_hi = round(mu + 1.96*se, 4)
        bg    = C['light_green'] if ri%2==0 else C['white']
        dist  = f"N({mu:.4f}, {s.get('var',0):.4f})"
        row_d = [name, dist, mu, s.get('var',0), sigma,
                 p_neg, p_pos, ic_lo, ic_hi]
        for ci, v in enumerate(row_d, 1):
            nf = '0.0000' if ci in (3,4,5,6,7,8,9) else None
            bold = ci == 1
            cell(ws5, ri, ci, v, bold=bold, bg=bg,
                 fg=COLORS_AX[ri-sr5-2] if ci==1 else '000000',
                 h='center', num_fmt=nf)

    # ── TCL ──
    tcl_row = sr5 + 9
    section_title(ws5, tcl_row, 1,
        '  TEOREMA CENTRAL DEL LÍMITE — Error estándar', span=9)
    tcl_hdrs = ['Variable','n','σ/√n (SE)','IC 95% Inf','IC 95% Sup',
                'Unidad','Interpretación','','']
    for ci, h in enumerate(tcl_hdrs, 1):
        hdr(ws5, tcl_row+1, ci, h)

    for ri, name in enumerate(AX, tcl_row+2):
        vals = ax_vals(name)
        s    = _stat(vals)
        mu, sigma = s.get('mean',0), s.get('std',0)
        n    = s.get('n',1) or 1
        se   = round(sigma/math.sqrt(n), 6)
        ic_lo= round(mu - 1.96*se, 4)
        ic_hi= round(mu + 1.96*se, 4)
        bg   = C['light_green'] if ri%2==0 else C['white']
        interp = (f'Con 95% de confianza, la media real de {name} '
                  f'está entre {ic_lo} y {ic_hi} {UNITS[name]}')
        for ci, v in enumerate([name, n, se, ic_lo, ic_hi,
                                 UNITS[name], interp,'',''], 1):
            nf = '0.0000' if ci in (3,4,5) else None
            cell(ws5, ri, ci, v, bold=(ci==1), bg=bg, h='center',
                 num_fmt=nf, wrap=(ci==7))
        ws5.merge_cells(start_row=ri, start_column=7,
                        end_row=ri, end_column=9)

    # ── Modelos discretos ──
    disc_row = tcl_row + 10
    section_title(ws5, disc_row, 1,
        '  MODELOS DISCRETOS — Binomial y Geométrica', span=9)
    disc_hdrs = ['Modelo','Variable','Evento','p','n','E[X]','σ','Interpretación','']
    for ci, h in enumerate(disc_hdrs, 1):
        hdr(ws5, disc_row+1, ci, h)

    # Binomial: Accel X > 0
    ax_v    = ax_vals('Accel X')
    p_bin   = round(sum(1 for v in ax_v if v>0)/len(ax_v), 4) if ax_v else 0
    n_bin   = len(ax_v)
    e_bin   = round(n_bin*p_bin, 2)
    s_bin   = round(math.sqrt(n_bin*p_bin*(1-p_bin)), 4) if p_bin not in (0,1) else 0
    interp_b = f'Se esperan {e_bin:.1f} disparos con Accel X > 0 de {n_bin} ensayos'
    for ci, v in enumerate(['Binomial','Accel X','Accel X > 0',
                             p_bin, n_bin, e_bin, s_bin, interp_b,''], 1):
        nf = '0.0000' if ci == 4 else ('0.00' if ci in (6,7) else None)
        cell(ws5, disc_row+2, ci, v, bg=C['light_green'],
             h='center', num_fmt=nf, wrap=(ci==8))

    # Geométrica: Gyro X < -30
    gx_v   = ax_vals('Gyro X')
    p_geo  = sum(1 for v in gx_v if v<-30)/len(gx_v) if gx_v else 0
    e_geo  = round(1/p_geo, 2) if p_geo > 0 else 'N/A (sin eventos en este dataset)'
    s_note = ('⚠ p=0 en este dataset — no hubo eventos extremos. '
              'Necesitas más disparos.' if p_geo==0 else
              f'Se esperan {e_geo} disparos hasta el 1er Gyro X < −30°/s')
    for ci, v in enumerate(['Geométrica','Gyro X','Gyro X < −30°/s',
                             round(p_geo,4) if p_geo>0 else 0,
                             '—', e_geo, '—', s_note,''], 1):
        nf = '0.0000' if ci == 4 else None
        cell(ws5, disc_row+3, ci, v, bg=C['white'],
             h='center', num_fmt=nf, wrap=(ci==8))

    # Nota final
    note5_row = disc_row + 5
    cell(ws5, note5_row, 1,
         'Nota: Los modelos son válidos con n ≥ 30 (TCL). '
         f'Dataset actual: n = {len(valid)}. '
         'Con pocos datos, los parámetros son estimados pero no robustos. '
         'Cada vez que descargues el Excel con más datos, los valores se recalculan automáticamente.',
         italic=True, fg=C['muted'], wrap=True, h='left',
         border=False, size=9)
    ws5.merge_cells(f'A{note5_row}:I{note5_row}')
    ws5.row_dimensions[note5_row].height = 35

    autowidth(ws5)

    # ═══════════════════════════════════════════════════════════════════════
    # HOJA 6 — GRÁFICAS DE EVOLUCIÓN POR DISPARO
    # ═══════════════════════════════════════════════════════════════════════
    ws6 = wb.create_sheet('Gráficas')
    section_title(ws6, 1, 1,
        '  EVOLUCIÓN POR DISPARO — Medias de cada eje', span=10)

    # Tabla de medias por disparo
    hdr(ws6, 2, 1, 'Disparo')
    for ci, name in enumerate(AX, 2):
        hdr(ws6, 2, ci, f'{name} ({UNITS[name]})')

    for ri, d in enumerate(d_ids, 3):
        bg = C['light_green'] if ri%2==0 else C['white']
        rs = by_d[d]
        means = [sum(getattr(r,k.lower().replace(' ','_'))
                     for r in rs)/len(rs) for k in AX]
        cell(ws6, ri, 1, d, bg=bg, h='center',
             fg=C['accent_blue'], bold=True)
        for ci, v in enumerate(means, 2):
            cell(ws6, ri, ci, round(v,4), bg=bg, h='center',
                 fg=C['positive'] if v>=0 else C['negative'],
                 num_fmt='0.0000')

    n_d = len(d_ids)
    if n_d >= 2:
        # Gráfico acelerómetro
        chart_accel = LineChart()
        chart_accel.title  = 'Acelerómetro por Disparo (g)'
        chart_accel.y_axis.title = 'g'
        chart_accel.x_axis.title = 'Número de disparo'
        chart_accel.width = 18; chart_accel.height = 12
        for ci, color in zip([2,3,4], ['4472C4','7030A0','70AD47']):
            ref = Reference(ws6, min_col=ci, min_row=2, max_row=2+n_d)
            chart_accel.add_data(ref, titles_from_data=True)
            chart_accel.series[-1].graphicalProperties.line.solidFill = color
        cats = Reference(ws6, min_col=1, min_row=3, max_row=2+n_d)
        chart_accel.set_categories(cats)
        ws6.add_chart(chart_accel, 'I3')

        # Gráfico giroscopio
        chart_gyro = LineChart()
        chart_gyro.title  = 'Giroscopio por Disparo (°/s)'
        chart_gyro.y_axis.title = '°/s'
        chart_gyro.x_axis.title = 'Número de disparo'
        chart_gyro.width = 18; chart_gyro.height = 12
        for ci, color in zip([5,6,7], ['FF0000','FFC000','FF6600']):
            ref = Reference(ws6, min_col=ci, min_row=2, max_row=2+n_d)
            chart_gyro.add_data(ref, titles_from_data=True)
            chart_gyro.series[-1].graphicalProperties.line.solidFill = color
        chart_gyro.set_categories(cats)
        ws6.add_chart(chart_gyro, 'I22')

    autowidth(ws6)

    # ─── Guardar ─────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()